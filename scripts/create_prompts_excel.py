import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ──────────────────────────────────────────
# BEDSHEET TAB
# ──────────────────────────────────────────
ws = wb.active
ws.title = "Bedsheet"

# Styles
header_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
subheader_font = Font(name="Calibri", size=11, bold=True, color="78350F")
subheader_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
body_font = Font(name="Calibri", size=11)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Column widths
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 90
ws.column_dimensions["C"].width = 45

# ── Title Row ──
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = "QuickLoom — Bedsheet Product Photography Prompts (Gemini 2.5 Flash)"
title_cell.font = Font(name="Calibri", size=15, bold=True, color="78350F")
title_cell.alignment = Alignment(vertical="center")
title_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
ws.row_dimensions[1].height = 38

# ── Instructions Row ──
ws.merge_cells("A2:C2")
instr = ws["A2"]
instr.value = (
    "HOW TO USE: Upload your raw bedsheet image to Gemini 2.5 Flash → Paste the prompt from column B → "
    "Generate the image. Use all 3 prompts per bedsheet to get a consistent set of Hero, Macro & Flat-Lay photos. "
    "Always upload the SAME raw image for all 3 prompts to ensure design/colour consistency."
)
instr.font = Font(name="Calibri", size=10, italic=True, color="92400E")
instr.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 48

# ── Blank spacer ──
ws.row_dimensions[3].height = 10

# ── Headers ──
headers = ["Image Type", "Gemini 2.5 Flash Prompt (Copy-Paste)", "Notes / Tips"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border
ws.row_dimensions[4].height = 32

# ── PROMPT 1: Hero Lifestyle Shot ──
row = 5
ws.row_dimensions[row].height = 340

ws.cell(row=row, column=1, value="📸 PHOTO 1\nHero Lifestyle Shot\n(Main Product Image)").font = subheader_font
ws.cell(row=row, column=1).fill = subheader_fill
ws.cell(row=row, column=1).alignment = center_alignment
ws.cell(row=row, column=1).border = thin_border

prompt1 = """You are a professional e-commerce product photographer specialising in Indian home textiles for middle-class families.

ROLE: You are creating the MAIN HERO product image for an online store called QuickLoom that sells affordable handloom bedsheets (starting ₹499) to families in Gurgaon & Bhiwadi. The target customer is a 28-45 year old Indian homemaker or couple who wants their bedroom to look beautiful without spending a fortune.

TASK: Using the uploaded raw bedsheet image as your EXACT design reference, generate a photorealistic 4K lifestyle product photograph following these STRICT rules:

DESIGN ACCURACY (CRITICAL):
— Extract the EXACT pattern, print, colour palette, motifs, and border design from the uploaded image
— The generated bedsheet MUST look identical to the uploaded image in terms of design, colours, and pattern placement
— Do NOT alter, simplify, or reimagine the design — reproduce it faithfully

SCENE SETUP:
— A clean, modern Indian middle-class bedroom (not luxury, not budget — relatable aspirational)
— Light wood or white headboard, simple bedside table with a small plant or lamp
— Warm, natural morning sunlight streaming through a window with sheer curtains
— The bedsheet is neatly spread on a queen-size bed with 2 matching pillows
— Subtle props: a chai cup on bedside table, a book, or reading glasses (homely Indian feel)
— Walls in soft neutral tone (off-white, warm beige, or pale sage)

CAMERA & LIGHTING:
— Shot from a 3/4 elevated angle (like standing at the foot of the bed looking slightly down)
— Golden hour warm lighting — makes the fabric colours look rich and inviting
— Soft shadows, no harsh highlights
— Depth of field: bed in sharp focus, background gently blurred
— Resolution: 4K, photorealistic, magazine-quality

MOOD: Warm, inviting, aspirational but ACHIEVABLE — the viewer should think "This would look perfect on MY bed" not "This is too fancy for me"

BACKGROUND: Clean, uncluttered. The bedsheet is the HERO — everything else is supporting.

OUTPUT: Single photorealistic image, landscape orientation (16:9 or 4:3), no text overlays, no watermarks, no borders."""

ws.cell(row=row, column=2, value=prompt1).font = body_font
ws.cell(row=row, column=2).alignment = wrap_alignment
ws.cell(row=row, column=2).border = thin_border

tips1 = """• This is the FIRST image customers see — it must trigger "I want this on my bed"
• Warm lighting makes ₹499 bedsheet look premium
• Indian middle-class bedroom = relatable
• Chai cup / book = subconscious "home" feeling
• Always use the SAME raw image for all 3 prompts
• If colours look off, add: "Match the exact hex/shade from reference"
• Landscape orientation works best for website hero cards"""

ws.cell(row=row, column=3, value=tips1).font = Font(name="Calibri", size=10, color="6B7280")
ws.cell(row=row, column=3).alignment = wrap_alignment
ws.cell(row=row, column=3).border = thin_border

# ── PROMPT 2: Macro Fabric Detail ──
row = 6
ws.row_dimensions[row].height = 320

ws.cell(row=row, column=1, value="🔍 PHOTO 2\nMacro Fabric Close-Up\n(Quality & Texture Shot)").font = subheader_font
ws.cell(row=row, column=1).fill = subheader_fill
ws.cell(row=row, column=1).alignment = center_alignment
ws.cell(row=row, column=1).border = thin_border

prompt2 = """You are a professional macro product photographer specialising in fabric and textile photography for Indian e-commerce.

ROLE: You are creating a MACRO CLOSE-UP shot for QuickLoom, an affordable handloom bedsheet brand. The target customer is a value-conscious Indian family who needs to SEE the fabric quality before trusting an online purchase. This image must answer: "Is the fabric actually good quality?"

TASK: Using the uploaded raw bedsheet image as your EXACT design reference, generate a photorealistic macro close-up photograph following these STRICT rules:

DESIGN ACCURACY (CRITICAL):
— The fabric pattern, weave, thread colours, and any embroidery/print visible in the uploaded image MUST be faithfully reproduced
— Show the REAL texture of the fabric — whether it's cotton weave, khadi texture, block print, or screen print
— Capture the individual thread structure if it's a handloom weave

COMPOSITION:
— Extreme close-up of the bedsheet fabric, filling 80% of the frame
— Show a section where the design/pattern meets a border or where stitching is visible
— Part of the fabric should be slightly folded or draped to show thickness and body
— Include the stitching/hem line if the design has one — show it is neat and even
— One corner of the frame can show the fabric edge/selvedge to indicate quality

CAMERA & LIGHTING:
— Macro lens feel — shallow depth of field with the weave pattern tack-sharp in the center
— Soft, diffused natural light from one side (like near a window)
— Light should graze the fabric surface to reveal texture and weave depth
— No harsh shadows — use a reflector feel from the opposite side
— Background: the same fabric going out of focus, or a clean wooden surface peeking at the edge
— Resolution: 4K, photorealistic

WHAT TO HIGHLIGHT:
— Thread count / weave density (shows quality)
— Colour vibrancy up close (not faded or dull)
— Stitching precision (straight, even, no loose threads)
— Fabric thickness / body (not see-through, not flimsy)

MOOD: Trustworthy, premium-feel-at-affordable-price. The viewer should think "This fabric looks solid, not cheap."

OUTPUT: Single photorealistic image, square or 4:5 portrait orientation, no text overlays, no watermarks."""

ws.cell(row=row, column=2, value=prompt2).font = body_font
ws.cell(row=row, column=2).alignment = wrap_alignment
ws.cell(row=row, column=2).border = thin_border

tips2 = """• This builds TRUST — middle-class buyers fear "cheap fabric"
• Side lighting reveals weave texture beautifully
• Folded edge shows thickness without words
• Neat stitching = quality signal for Indian buyers
• If fabric is handloom, ask Gemini to emphasise the handwoven texture
• Square format works well as 2nd image in product gallery
• Add "Show the cotton weave structure" if it's a cotton bedsheet"""

ws.cell(row=row, column=3, value=tips2).font = Font(name="Calibri", size=10, color="6B7280")
ws.cell(row=row, column=3).alignment = wrap_alignment
ws.cell(row=row, column=3).border = thin_border

# ── PROMPT 3: Flat Lay / Full Design ──
row = 7
ws.row_dimensions[row].height = 320

ws.cell(row=row, column=1, value="🎨 PHOTO 3\nFlat-Lay Full Design\n(Pattern & Colour Accuracy)").font = subheader_font
ws.cell(row=row, column=1).fill = subheader_fill
ws.cell(row=row, column=1).alignment = center_alignment
ws.cell(row=row, column=1).border = thin_border

prompt3 = """You are a professional flat-lay product photographer specialising in Indian textile e-commerce photography.

ROLE: You are creating a FLAT-LAY / FULL DESIGN shot for QuickLoom, an affordable handloom bedsheet brand for middle-class Indian families. This image must show the COMPLETE design, pattern, and colour of the bedsheet so the customer knows EXACTLY what they are ordering. No surprises.

TASK: Using the uploaded raw bedsheet image as your EXACT design reference, generate a photorealistic flat-lay photograph following these STRICT rules:

DESIGN ACCURACY (CRITICAL):
— Reproduce the COMPLETE pattern from the uploaded image — full motifs, borders, corner designs, centre pattern
— Colours must match the uploaded image EXACTLY — same saturation, same tones
— If the bedsheet has a border pattern, show ALL four borders clearly
— Print/embroidery placement must match the original

COMPOSITION:
— Top-down flat-lay shot — camera directly above, looking straight down
— The bedsheet is laid flat on a clean, contrasting surface (light wooden floor, marble, or off-white fabric)
— One corner is artfully folded back to reveal the reverse side / underside texture
— The ENTIRE design is visible — borders, centre pattern, pillow cover designs if included
— Include the 2 matching pillow covers arranged neatly on top or beside the bedsheet
— Subtle styling props placed at the edges: a small potted plant, dried flowers, or a decorative tray (minimal, not distracting)

CAMERA & LIGHTING:
— Perfect top-down angle (bird's eye view / 90°)
— Even, diffused overhead lighting — no shadows, no hot spots
— Colours should appear TRUE TO LIFE — exactly as they would look in a well-lit room
— White balance: neutral warm — not too cool, not too yellow
— Resolution: 4K, photorealistic, clean e-commerce quality

WHAT TO HIGHLIGHT:
— Complete design visibility (customer sees the full pattern before buying)
— Colour accuracy (what you see is what you get)
— Set completeness (bedsheet + pillow covers)
— Fabric smoothness and quality of print/weave

MOOD: Clear, honest, what-you-see-is-what-you-get. The viewer should think "Now I know exactly what this looks like — I'm confident ordering this."

OUTPUT: Single photorealistic image, square (1:1) orientation, no text overlays, no watermarks, no borders. Clean white/light background at edges."""

ws.cell(row=row, column=2, value=prompt3).font = body_font
ws.cell(row=row, column=2).alignment = wrap_alignment
ws.cell(row=row, column=2).border = thin_border

tips3 = """• This is the "confidence" image — removes buyer doubt
• Folded corner = shows both sides = transparency
• Include pillow covers to show "complete set" value
• Even lighting = true colour representation
• Middle-class buyers compare designs carefully — show the FULL pattern
• Square format for consistent product gallery
• If bedsheet has elastic, show it in the folded corner
• This image + Hero image = most viewed pair in product pages"""

ws.cell(row=row, column=3, value=tips3).font = Font(name="Calibri", size=10, color="6B7280")
ws.cell(row=row, column=3).alignment = wrap_alignment
ws.cell(row=row, column=3).border = thin_border

# ── Consistency Note Row ──
row = 9
ws.merge_cells("A9:C9")
note = ws["A9"]
note.value = (
    "⚠️ CONSISTENCY RULE: For every new bedsheet, upload the SAME raw image to all 3 prompts. "
    "This ensures the design, colours, and pattern are identical across Hero, Macro, and Flat-Lay shots. "
    "Never mix raw images between prompts for the same product."
)
note.font = Font(name="Calibri", size=11, bold=True, color="DC2626")
note.alignment = Alignment(wrap_text=True, vertical="center")
note.fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
ws.row_dimensions[9].height = 52

# ── Quick Reference Row ──
row = 11
ws.merge_cells("A11:C11")
ref = ws["A11"]
ref.value = "QUICK REFERENCE — Image Specs per Product"
ref.font = Font(name="Calibri", size=12, bold=True, color="78350F")
ref.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
ws.row_dimensions[11].height = 30

specs = [
    ["Photo 1 — Hero", "Landscape 16:9 or 4:3", "4K", "Lifestyle on bed", "Main product card, hero banner"],
    ["Photo 2 — Macro", "Square 1:1 or Portrait 4:5", "4K", "Extreme close-up", "2nd image in gallery, builds trust"],
    ["Photo 3 — Flat-Lay", "Square 1:1", "4K", "Top-down full pattern", "3rd image, shows complete design"],
]

spec_headers = ["Image", "Orientation", "Resolution", "Shot Type", "Used For"]
for col_idx, h in enumerate(spec_headers, 1):
    if col_idx <= 3:
        cell = ws.cell(row=12, column=col_idx, value=h)
    # We need 5 columns for specs, let's adjust
# Actually let's put specs in A-C as a formatted table
ws.cell(row=12, column=1, value="Image").font = Font(name="Calibri", size=10, bold=True)
ws.cell(row=12, column=2, value="Orientation / Resolution / Shot Type").font = Font(name="Calibri", size=10, bold=True)
ws.cell(row=12, column=3, value="Used For").font = Font(name="Calibri", size=10, bold=True)
for cell in [ws.cell(row=12, column=c) for c in range(1,4)]:
    cell.border = thin_border
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

for i, spec in enumerate(specs):
    r = 13 + i
    ws.cell(row=r, column=1, value=spec[0]).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=2, value=f"{spec[1]}  |  {spec[2]}  |  {spec[3]}").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=2).border = thin_border
    ws.cell(row=r, column=3, value=spec[4]).font = Font(name="Calibri", size=10, color="6B7280")
    ws.cell(row=r, column=3).border = thin_border

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "QuickLoom_Product_Photo_Prompts.xlsx")
output_path = os.path.normpath(output_path)
wb.save(output_path)
print(f"✅ Excel saved at: {output_path}")
